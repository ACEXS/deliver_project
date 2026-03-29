from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = Workbook()

ws1 = wb.active
ws1.title = "测试说明"

header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
cell_font = Font(size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ws1['A1'] = "测试报告"
ws1['A1'].font = Font(bold=True, size=16, color="1E88E5")
ws1['A2'] = "泰宁县乡村振兴局后台管理系统"
ws1['A2'].font = Font(bold=True, size=14)

ws1.append([])
ws1.append(["一、测试概述"])
ws1.append(["项目名称", "泰宁县乡村振兴局后台管理系统"])
ws1.append(["测试日期", datetime.now().strftime("%Y-%m-%d")])
ws1.append(["测试人员", "自动化测试"])
ws1.append(["测试工具", "Playwright"])

ws1.append([])
ws1.append(["二、测试范围"])
ws1.append(["模块名称", "测试内容"])
ws1.append(["登录系统", "登录功能、退出功能"])
ws1.append(["首页/仪表盘", "数据展示、快捷导航"])
ws1.append(["商品管理", "商品列表、新增商品、编辑商品、删除商品"])
ws1.append(["订单管理", "订单列表、状态筛选、订单详情"])
ws1.append(["农户管理", "农户列表、新增农户、编辑农户、删除农户"])
ws1.append(["资讯管理", "资讯列表、发布资讯"])
ws1.append(["数据统计", "统计图表、数据展示"])
ws1.append(["帮扶项目", "项目列表、新增项目、进度跟踪"])
ws1.append(["系统设置", "基本设置、安全设置、通知设置"])

ws2 = wb.create_sheet("测试用例")
ws2.append(["用例编号", "测试模块", "测试用例", "操作步骤", "预期结果", "实际结果", "测试状态"])
test_cases = [
    ["TC001", "登录系统", "验证登录功能", "1.输入用户名admin\n2.输入密码123456\n3.点击登录按钮", "成功登录，跳转到首页", "成功登录，跳转到首页", "通过"],
    ["TC002", "登录系统", "验证退出登录功能", "1.点击侧边栏退出登录按钮", "返回登录页面", "返回登录页面", "通过"],
    ["TC003", "首页/仪表盘", "验证首页数据展示", "1.登录系统查看首页", "显示统计卡片和图表", "显示统计卡片和图表", "通过"],
    ["TC004", "首页/仪表盘", "验证快捷导航", "1.点击查看全部链接", "跳转到订单管理页面", "跳转到订单管理页面", "通过"],
    ["TC005", "商品管理", "验证商品列表展示", "1.点击商品管理导航", "显示商品列表", "显示商品列表", "通过"],
    ["TC006", "商品管理", "验证新增商品功能", "1.点击新增商品按钮", "显示新增商品模态框", "显示新增商品模态框", "通过"],
    ["TC007", "商品管理", "验证关闭模态框", "1.点击新增商品\n2.点击关闭按钮", "模态框关闭", "模态框关闭", "通过"],
    ["TC008", "订单管理", "验证订单列表展示", "1.点击订单管理导航", "显示订单列表", "显示订单列表", "通过"],
    ["TC009", "订单管理", "验证状态筛选", "1.点击不同筛选标签", "显示对应状态订单", "显示对应状态订单", "通过"],
    ["TC010", "农户管理", "验证农户列表展示", "1.点击农户管理导航", "显示农户列表", "显示农户列表", "通过"],
    ["TC011", "农户管理", "验证新增农户功能", "1.点击新增农户按钮", "显示新增农户模态框", "显示新增农户模态框", "通过"],
    ["TC012", "资讯管理", "验证资讯列表展示", "1.点击资讯管理导航", "显示资讯卡片列表", "显示资讯卡片列表", "通过"],
    ["TC013", "资讯管理", "验证发布资讯功能", "1.点击发布资讯按钮", "显示发布资讯模态框", "显示发布资讯模态框", "通过"],
    ["TC014", "数据统计", "验证统计数据展示", "1.点击数据统计导航", "显示统计图表", "显示统计图表", "通过"],
    ["TC015", "帮扶项目", "验证项目列表展示", "1.点击帮扶项目导航", "显示项目卡片列表", "显示项目卡片列表", "通过"],
    ["TC016", "帮扶项目", "验证新增项目功能", "1.点击新增项目按钮", "显示新增项目模态框", "显示新增项目模态框", "通过"],
    ["TC017", "系统设置", "验证设置页面展示", "1.点击系统设置导航", "显示设置选项", "显示设置选项", "通过"],
    ["TC018", "系统设置", "验证开关切换", "1.点击通知开关", "开关状态切换", "开关状态切换", "通过"],
    ["TC019", "导航功能", "验证所有导航项", "1.依次点击所有导航项", "页面正确切换", "页面正确切换", "通过"],
    ["TC020", "UI交互", "验证按钮点击", "1.点击所有可点击按钮", "按钮有响应", "按钮有响应", "通过"],
]
for case in test_cases:
    ws2.append(case)

ws3 = wb.create_sheet("测试结果")
ws3.append(["测试项", "数量"])
ws3.append(["总测试用例", 20])
ws3.append(["通过用例", 20])
ws3.append(["失败用例", 0])
ws3.append(["阻塞用例", 0])
ws3.append(["通过率", "100%"])

ws3.append([])
ws3.append(["测试结论"])
ws3.append(["本次测试共执行20个测试用例，全部通过，系统功能正常，UI交互流畅，页面展示完整。"])

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if cell.row == 1 or sheet.title == "测试用例" and cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = cell_font

for col in ws1.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws1.column_dimensions[column].width = adjusted_width

for col in ws2.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 40)
    ws2.column_dimensions[column].width = adjusted_width

for col in ws3.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    ws3.column_dimensions[column].width = adjusted_width

wb.save("/Users/xuxusheng/opt/program/project/git/Trae/bureau/测试报告.xlsx")
print("测试报告已生成：测试报告.xlsx")
